from file_service.test_file.test_file_path import get_test_file_path
from file_service.report_all import get_report_file_all_path
from utils.db_api.core import DatabaseService1, Question
from file_service.hisobot import get_report_file_path
from file_service.file_path import get_file_path
from LoggingService import LoggerService
from openpyxl import load_workbook, Workbook
from datetime import datetime
from os.path import dirname
import subprocess

db = DatabaseService1(logger=LoggerService())
YO_NALISHLAR_QATORLARI = {
    "Inson resurslarini boshqarish": 5,
    "Hayot faoliyati xavfsizligi": 6,
    "Yurisprudensiya": 7,
    "Ijtimoiy ish": 8,
    "Menejment": 9,
    "Mehnat muhofazasi va texnika xavfsizligi": 10,
    "Psixologiya": 11,
    "Bugalteriya hisobi": 12,
    "Metrologiya va standartlashtirish": 13
}

async def create_report_by_faculty_files(rows: list) -> list[str]:
    file_paths = []
    faculties = set(row["faculty"] for row in rows)
    now = datetime.now()
    formatted_time = now.strftime("%d-%m-%Y")

    for faculty in faculties:
        faculty_rows = [row for row in rows if row["faculty"] == faculty]
        if not faculty_rows:
            continue

        template_path = await get_file_path("report_all.xlsx")
        workbook = load_workbook(template_path)
        sheet = workbook.active
        sheet["T2"] = f"{formatted_time} holatiga {faculty} yo‘nalishidagi abituriyentlar natijalari"

        start_row = 5
        for i, row in enumerate(faculty_rows):
            row_index = start_row + i
            sheet.cell(row=row_index, column=1, value=i + 1)
            sheet.cell(row=row_index, column=2, value=row["exam_day"])
            sheet.cell(row=row_index, column=3, value=row["name"])
            sheet.cell(row=row_index, column=4, value=row["faculty"])
            sheet.cell(row=row_index, column=5, value=row["total_ball"])

            for block in [1, 2, 3]:
                base_col = 6 + (block - 1) * 7
                sheet.cell(row=row_index, column=base_col + 0, value=row[f"block_{block}_jami_ball"])
                sheet.cell(row=row_index, column=base_col + 1, value=row[f"block_{block}_ball"])
                sheet.cell(row=row_index, column=base_col + 2, value=row[f"block_{block}_total_questions"])
                sheet.cell(row=row_index, column=base_col + 3, value=row[f"block_{block}_correct"])
                sheet.cell(row=row_index, column=base_col + 4, value=row[f"block_{block}_wrong"])
                sheet.cell(row=row_index, column=base_col + 5, value=row[f"block_{block}_accuracy"])
                sheet.cell(row=row_index, column=base_col + 6, value=row[f"block_{block}_subject"])

        file_name = f"{faculty.replace(' ', '_')}_{formatted_time}.xlsx"
        save_path = await get_report_file_all_path(file_name)
        workbook.save(save_path)
        workbook.close()
        file_paths.append(save_path)

    return file_paths



async def create_report_all_file(rows: list, status: str) -> str:
    """
    Abituriyentlar test natijalarini Excel faylga aniq va siljimasdan yozish funksiyasi.
    :param rows: get_report_handler orqali kelgan natijaviy ro'yxat
    :return: Excel faylning saqlangan to‘liq yo‘li
    """
    try:
        file_name = ""
        # 1. Shablonni ochamiz
        template_path = await get_file_path("report_all.xlsx")
        workbook = load_workbook(template_path)
        sheet = workbook.active
        now = datetime.now()
        formatted_time = now.strftime("%d-%m-%Y")
        if status == "all":
            file_name = now.strftime("%d-%m-%Y") + "_hisobot.xlsx"
            sheet["T2"] = f"{formatted_time} Holatiga imtihon toshirgan abituriyentlar natijalari"
        elif status == "daily":
            file_name = now.strftime("%d-%m-%Y") + "_kunlik_hisobot.xlsx"
            sheet["T2"] = f"{formatted_time} kunlik imtihon toshirgan abituriyentlar natijalari"
        start_row = 5  # Ma'lumotlar yozilishi kerak bo‘lgan qatordan boshlaymiz (MERGE bilan to‘qnashmaydi)

        for i, row in enumerate(rows):
            row_index = start_row + i

            sheet.cell(row=row_index, column=1, value=i + 1)

            # A–D ustunlar
            sheet.cell(row=row_index, column=2, value=row["exam_day"])        # A
            sheet.cell(row=row_index, column=3, value=row["name"])            # B
            sheet.cell(row=row_index, column=4, value=row["faculty"])         # C
            sheet.cell(row=row_index, column=5, value=row["total_ball"])      # D

            for block in [1, 2, 3]:
                base_col = 6 + (block - 1) * 7  # E=5, L=12, S=19
                sheet.cell(row=row_index, column=base_col + 0, value=row[f"block_{block}_jami_ball"])
                sheet.cell(row=row_index, column=base_col + 1, value=row[f"block_{block}_ball"])
                sheet.cell(row=row_index, column=base_col + 2, value=row[f"block_{block}_total_questions"])
                sheet.cell(row=row_index, column=base_col + 3, value=row[f"block_{block}_correct"])
                sheet.cell(row=row_index, column=base_col + 4, value=row[f"block_{block}_wrong"])
                sheet.cell(row=row_index, column=base_col + 5, value=row[f"block_{block}_accuracy"])
                sheet.cell(row=row_index, column=base_col + 6, value=row[f"block_{block}_subject"])
        save_path = await get_report_file_all_path(file_name)
        workbook.save(save_path)
        workbook.close()
        return save_path
    except Exception as e:
        print(f"❌ Excel faylga yozishda xatolik: {e}")
        raise


async def create_report_file(jami_data: dict, kunlik_data: dict, exam_data: dict, exam_all_data: dict) -> str:
    path = await get_file_path("report_finaly.xlsx")
    workbook = load_workbook(path)
    sheet = workbook.active
    total_daily = 0
    total_all = 0
    total_exam = 0
    total_exam_all = 0
    for yo_nalish, row in YO_NALISHLAR_QATORLARI.items():
        # Jami hujjat topshirganlar → E ustun
        jami_val = jami_data.get(yo_nalish, 0)
        sheet[f"E{row}"] = jami_val
        total_all += jami_val

        # Kunlik topshirganlar → D ustun
        kunlik_val = kunlik_data.get(yo_nalish, 0)
        sheet[f"D{row}"] = kunlik_val
        total_daily += kunlik_val

        # Imtihon topshirganlar → F ustun
        exam_val = exam_data.get(yo_nalish, 0)
        sheet[f"F{row}"] = exam_val
        total_exam += exam_val

        # Imtihon topshirganlar (barcha) → G ustun
        exam_all_val = exam_all_data.get(yo_nalish, 0)
        sheet[f"G{row}"] = exam_all_val
        total_exam_all += exam_all_val

    now = datetime.now()
    formatted_time = now.strftime("%d-%m-%Y")
    sheet["G2"] = f"{formatted_time} gacha holat bo'yicha"
    sheet["D4"] = formatted_time
    sheet["F4"] = formatted_time
    sheet["H4"] = formatted_time

    new_file_name = f"{formatted_time}_hisobot.xlsx"
    new_file_path = await get_report_file_path(new_file_name)

    workbook.save(new_file_path)
    workbook.close()
    return True


async def convert_pdf(source_path, target_path):
    source_path1 = await get_report_file_path(source_path)
    target_dir = dirname(source_path1)

    try:
        subprocess.run([
            'libreoffice', '--headless', '--convert-to', 'pdf', source_path1, '--outdir', target_dir
        ], check=True)

        print("✅ PDF muvaffaqiyatli yaratildi.")
    except subprocess.CalledProcessError as e:
        print(f"❌ LibreOffice orqali PDFga o‘tkazishda xatolik: {e}")


async def write_qabul(row: list):
    try:
        path = await get_file_path("qabul.xlsx")
        workbook = load_workbook(path)
        sheet = workbook.active
        sheet.append(row)  # to'g'ridan-to'g'ri ro'yxat yoziladi
        workbook.save(path)
        return True
    except Exception as e:
        print(f"❌ Excel yozishda xatolik: {e}")
        return False


async def read_file(file_path: str, subject_id: int) -> bool:
    """
    Excel fayldagi savollarni o'qiydi va bazaga qo'shadi.
    - file_path: yuklangan fayl nomi (path olish uchun get_test_file_path ishlatiladi)
    - subject_id: savollar tegishli bo'lgan fan ID
    Qaytaradi True agar yangi savollar qo'shilgan bo'lsa, False aks holda.
    """
    try:
        # 1) Faylni yuklab olish va Workbook ochish
        full_path = await get_test_file_path(name=file_path)
        wb = load_workbook(full_path)
        sheet = wb.active

        # 2) 2-qatordan boshlab barcha satrlarni olish
        rows = list(sheet.iter_rows(min_row=1, values_only=True))
        if not rows:
            print("Faylda savollar topilmadi.")
            return False

        # 3) Birinchi ma'lumot qatori bo'yicha baza tekshiruvi
        first = rows[0]

        if len(first) < 2 or first[1] is None:
            print("Fayl formati noto‘g‘ri yoki birinchi savol yo‘q.")
            return False
        first_text = str(first[1]).strip()
        print(first_text)
        exists = await db.get(Question, filters={"subject_id": subject_id, "text": first_text})
        if exists:
            print("Ma'lumotlar bazasida allaqachon bu savol mavjud:", first_text)
            return False

        # 4) Har bir satr bo‘yicha yangi Question yaratish
        added = 0
        for row in rows:
            # row: tuple (A2, B2, C2, D2, E2, F2, …)
            if not row or len(row) < 1:
                # kam ma'lumotli satrlarni o‘tkazib yuboramiz
                continue
            print(row[0], row[1], row[2], row[3], row[4])

            text = str(row[0] or "").strip()
            opt1 = str(row[1] or "").strip()
            opt2 = str(row[2] or "").strip()
            opt3 = str(row[3] or "").strip()
            opt4 = str(row[4] or "").strip()
            # Misolda to‘g‘ri javob 3-ustunda ekan: row[2]
            correct = opt1

            if not text:
                continue

            await db.add(Question(
                subject_id=subject_id,
                text=text,
                option1=opt1,
                option2=opt2,
                option3=opt3,
                option4=opt4,
                correct_answer=correct
            ))
            added += 1

        print(f"{added} ta savol bazaga qo‘shildi.")
        return added > 0

    except Exception as e:
        print("read_file xatosi:", e)
        return False


async def check_passport_exists(passport_seria: str) -> bool:
    try:
        path = await get_file_path("student_db.xlsx")
        workbook = load_workbook(path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # boshqatorni o'tkazib yuboramiz
            if row[2] == passport_seria:
                return True

        return False

    except Exception as e:
        print(f"❌ Excel yozishda xatolik: {e}")
        return False
